from flask import Flask, jsonify, request, url_for
from flask_cors import CORS
from openpyxl import load_workbook

app = Flask(__name__)
CORS(app)

EXCEL_FILE = "./dataset/register.xlsx"

def login_func():
    data = request.get_json()
    email = data.get("email")
    password = data.get("password")

    try:
        wb = load_workbook(EXCEL_FILE)
        sheet = wb.active

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if len(row) >= 3:
                stored_email = row[0]
                stored_password = row[2]

                if email == stored_email and password == stored_password:
                    return jsonify({
                        "message": "Login successful!",
                        "redirect": url_for('personal', email=email)
                    }), 200

        return jsonify({"error": "Invalid email or password!"}), 401

    except Exception as e:
        return jsonify({"error": str(e)}), 500

if __name__ == '__main__':
    app.run(debug=True)
