from flask import Flask, jsonify, request, url_for
from flask_cors import CORS
from openpyxl import load_workbook
import pandas as pd

app = Flask(__name__)
CORS(app)  # allow React (localhost:3000/5173) to hit these APIs

# Excel files
REGISTER_XLSX = "./dataset/register.xlsx"
CHOOSE_XLSX   = "./dataset/choose.xlsx"

# Preload college info
college_data = pd.read_excel(CHOOSE_XLSX)
college_info = {}
for _, row in college_data.iterrows():
    c = row['College Name']
    loc = row['Location']
    courses = [
        x.strip('- ').strip()
        for x in str(row['Courses Offered']).split('\n')
        if x.strip()
    ]
    college_info[c] = {"location": loc, "courses": courses}


def get_colleges_func():
    return jsonify(list(college_info.keys()))


def get_details_func():
    payload = request.get_json() or {}
    c = payload.get("college", "")
    return jsonify(college_info.get(c, {"location": "", "courses": []}))


def submit_func():
    data = request.get_json() or {}
    email = data.get("email")
    college = data.get("college")
    course = data.get("course")
    location = data.get("location")

    try:
        wb = load_workbook(REGISTER_XLSX)
        sheet = wb.active

        # Find the user’s row by their email
        for row_idx in range(2, sheet.max_row + 1):
            if sheet.cell(row=row_idx, column=1).value == email:
                # Columns D=4, E=5, F=6, G=7, H=8, I=9
                for col_idx, val in zip([4, 5, 6, 7, 8, 9],
                                        [college, course, location,
                                         college, course, location]):
                    sheet.cell(row=row_idx, column=col_idx, value=val)

                wb.save(REGISTER_XLSX)
                return jsonify({
                    "message": "Data saved successfully!",
                    "redirect": url_for('company')  # or wherever
                }), 200

        return jsonify({"error": "User not found!"}), 404

    except Exception as e:
        return jsonify({"error": str(e)}), 500


if __name__ == "__main__":
    app.run(port=5000, debug=True)
